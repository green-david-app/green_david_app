
import os, re, io, base64, sqlite3
from datetime import datetime
from flask import Flask, send_from_directory, request, jsonify, session, g, send_file, abort
from werkzeug.security import generate_password_hash, check_password_hash

DB_PATH = os.environ.get("DB_PATH", "app.db")
SECRET_KEY = os.environ.get("SECRET_KEY", "dev-" + os.urandom(16).hex())
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "uploads")

app = Flask(__name__, static_folder=".", static_url_path="")
app.secret_key = SECRET_KEY
os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()

# ---------- migrations ----------
def get_version(db):
    db.execute("CREATE TABLE IF NOT EXISTS app_meta (id INTEGER PRIMARY KEY CHECK (id=1), version INTEGER NOT NULL)")
    row = db.execute("SELECT version FROM app_meta WHERE id=1").fetchone()
    if row is None:
        db.execute("INSERT INTO app_meta(id,version) VALUES (1,0)")
        db.commit()
        return 0
    return row["version"]

def set_version(db, v):
    db.execute("UPDATE app_meta SET version=? WHERE id=1", (v,))
    db.commit()

def migrate(db):
    v = get_version(db)
    if v < 1:
        c = db.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','manager','worker')),
            password_hash TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            client TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Plán',
            city TEXT NOT NULL,
            code TEXT NOT NULL,
            date TEXT NOT NULL,
            note TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            role TEXT NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site TEXT NOT NULL CHECK(site IN ('lipnik','praha')),
            category TEXT NOT NULL,
            name TEXT NOT NULL,
            qty REAL NOT NULL DEFAULT 0,
            unit TEXT NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS job_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            qty REAL NOT NULL,
            unit TEXT NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS job_tools (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            qty REAL NOT NULL,
            unit TEXT NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS job_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            created_at TEXT NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS job_assignments (
            job_id INTEGER NOT NULL,
            employee_id INTEGER NOT NULL,
            PRIMARY KEY(job_id, employee_id)
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS timesheets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            job_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            hours REAL NOT NULL,
            place TEXT,
            activity TEXT
        )""")
        db.commit()
        set_version(db, 1)
        v = 1
    if v < 2:
        db.execute("""CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            status TEXT NOT NULL DEFAULT 'nový',
            due_date TEXT,
            employee_id INTEGER,
            job_id INTEGER,
            created_by INTEGER,
            created_at TEXT NOT NULL
        )""")
        db.commit()
        set_version(db, 2)

def seed_admin(db):
    cur = db.execute("SELECT COUNT(*) as c FROM users")
    if cur.fetchone()["c"] == 0:
        db.execute("""INSERT INTO users(email,name,role,password_hash,active,created_at)
                      VALUES (?,?,?,?,1,?)""",
                   (os.environ.get("ADMIN_EMAIL","admin@greendavid.local"),
                    os.environ.get("ADMIN_NAME","Admin"),
                    "admin",
                    generate_password_hash(os.environ.get("ADMIN_PASSWORD","admin123")),
                    datetime.utcnow().isoformat()))
        db.commit()

@app.before_request
def ensure_db():
    db = get_db()
    migrate(db)
    seed_admin(db)

# ---------- static ----------
@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/uploads/<path:name>")
def uploaded_file(name):
    safe = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
    path = os.path.join(UPLOAD_DIR, safe)
    if not os.path.isfile(path): abort(404)
    return send_from_directory(UPLOAD_DIR, safe)

@app.route("/health")
def health():
    return {"status": "ok"}

# ---------- auth helpers ----------
def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    db = get_db()
    row = db.execute("SELECT id,email,name,role,active FROM users WHERE id=?",(uid,)).fetchone()
    return dict(row) if row else None

def require_auth():
    u = current_user()
    if not u or not u.get("active"):
        return None, (jsonify({"ok": False, "error": "unauthorized"}), 401)
    return u, None

def require_role(write=False):
    u, err = require_auth()
    if err: return None, err
    if write and u["role"] not in ("admin","manager"):
        return None, (jsonify({"ok": False, "error": "forbidden"}), 403)
    return u, None

# ---------- auth ----------
@app.route("/api/me")
def api_me():
    u = current_user()
    tasks_count = 0
    if u:
        db = get_db()
        tasks_count = db.execute("SELECT COUNT(*) c FROM tasks WHERE employee_id=? AND status!='hotovo'", (u["id"],)).fetchone()["c"]
    return jsonify({"ok": True, "authenticated": bool(u), "user": u, "tasks_count": tasks_count})

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    db = get_db()
    row = db.execute("SELECT id,email,name,role,password_hash,active FROM users WHERE email=?", (email,)).fetchone()
    if not row or not check_password_hash(row["password_hash"], password) or not row["active"]:
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401
    session["uid"] = row["id"]
    return jsonify({"ok": True})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.pop("uid", None)
    return jsonify({"ok": True})

# ---------- users (admin) ----------
@app.route("/api/users", methods=["GET","POST","PATCH"])
def api_users():
    u, err = require_auth()
    if err: return err
    db = get_db()
    if request.method == "GET":
        if u["role"] != "admin":
            return jsonify({"ok": False, "error": "forbidden"}), 403
        rows = db.execute("SELECT id,email,name,role,active,created_at FROM users ORDER BY id").fetchall()
        return jsonify({"ok": True, "users":[dict(r) for r in rows]})
    if u["role"] != "admin":
        return jsonify({"ok": False, "error": "forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    if request.method == "POST":
        email = (data.get("email") or "").strip().lower()
        name = data.get("name") or ""
        role = data.get("role") or "worker"
        pwd  = data.get("password") or ""
        if not (email and name and pwd and role in ("admin","manager","worker")):
            return jsonify({"ok": False, "error":"invalid_input"}), 400
        try:
            db.execute("""INSERT INTO users(email,name,role,password_hash,active,created_at)
                          VALUES (?,?,?,?,1,?)""",
                       (email, name, role, generate_password_hash(pwd), datetime.utcnow().isoformat()))
            db.commit()
            return jsonify({"ok": True})
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error":"email_exists"}), 400
    if request.method == "PATCH":
        uid = data.get("id")
        updates = []; params = []
        if "role" in data:
            if data["role"] not in ("admin","manager","worker"):
                return jsonify({"ok": False, "error":"bad_role"}), 400
            updates.append("role=?"); params.append(data["role"])
        if "active" in data:
            updates.append("active=?"); params.append(1 if data["active"] else 0)
        if "password" in data and data["password"]:
            updates.append("password_hash=?"); params.append(generate_password_hash(data["password"]))
        if not uid or not updates:
            return jsonify({"ok": False, "error":"invalid_input"}), 400
        params.append(uid)
        db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
        return jsonify({"ok": True})

# ---------- employees + timesheets ----------
@app.route("/api/employees", methods=["GET","POST","DELETE"])
def api_employees():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()
    if request.method == "GET":
        rows = db.execute("SELECT * FROM employees ORDER BY id DESC").fetchall()
        return jsonify({"ok": True, "employees":[dict(r) for r in rows]})
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        name = data.get("name"); role = data.get("role")
        if not (name and role): return jsonify({"ok": False, "error":"invalid_input"}), 400
        db.execute("INSERT INTO employees(name,role) VALUES (?,?)", (name, role))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        eid = request.args.get("id", type=int)
        if not eid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM employees WHERE id=?", (eid,))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/timesheets", methods=["GET","POST","DELETE"])
def api_timesheets():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()
    if request.method == "GET":
        emp = request.args.get("employee_id", type=int)
        jid = request.args.get("job_id", type=int)
        q = """SELECT t.id,t.employee_id,t.job_id,t.date,t.hours,t.place,t.activity,
                      e.name AS employee_name, j.title AS job_title, j.code AS job_code
               FROM timesheets t
               LEFT JOIN employees e ON e.id=t.employee_id
               LEFT JOIN jobs j ON j.id=t.job_id"""
        conds=[]; params=[]
        if emp: conds.append("t.employee_id=?"); params.append(emp)
        if jid: conds.append("t.job_id=?"); params.append(jid)
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY t.date DESC, t.id DESC"
        rows = db.execute(q, params).fetchall()
        return jsonify({"ok": True, "rows":[dict(r) for r in rows]})
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        emp = data.get("employee_id"); job = data.get("job_id"); date=data.get("date"); hours = data.get("hours")
        place = data.get("place") or ""
        activity = data.get("activity") or ""
        if not all([emp,job,date,hours is not None]): return jsonify({"ok": False, "error":"invalid_input"}), 400
        db.execute("INSERT INTO timesheets(employee_id,job_id,date,hours,place,activity) VALUES (?,?,?,?,?,?)", (emp,job,date,float(hours),place,activity))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        tid = request.args.get("id", type=int)
        if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM timesheets WHERE id=?", (tid,))
        db.commit()
        return jsonify({"ok": True})

# ---------- warehouse ----------
VALID_CATS = ('trvalky','trávy','dřeviny','stromy','cibuloviny','hnojiva/postřiky','materiál zahrada','materiál stavba')

@app.route("/api/items", methods=["GET","POST","PATCH","DELETE"])
def api_items():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()
    if request.method == "GET":
        site = request.args.get("site")
        q = "SELECT * FROM items" + (" WHERE site=?" if site in ('lipnik','praha') else "") + " ORDER BY id DESC"
        rows = db.execute(q, (site, ) if site in ('lipnik','praha') else ()).fetchall()
        return jsonify({"ok": True, "items":[dict(r) for r in rows]})
    data = request.get_json(force=True, silent=True) or {}
    if request.method == "POST":
        req = ["site","category","name","qty","unit"]
        if not all(k in data and data[k] not in (None,"") for k in req):
            return jsonify({"ok": False, "error":"invalid_input"}), 400
        if data["site"] not in ("lipnik","praha") or data["category"] not in VALID_CATS:
            return jsonify({"ok": False, "error":"bad_site_or_category"}), 400
        db.execute("""INSERT INTO items(site,category,name,qty,unit) VALUES (?,?,?,?,?)""",
                   (data["site"], data["category"], data["name"], float(data["qty"]), data["unit"]))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "PATCH":
        iid = data.get("id")
        if not iid: return jsonify({"ok": False, "error":"missing_id"}), 400
        updates=[]; params=[]
        for f in ["site","category","name","qty","unit"]:
            if f in data:
                if f=="site" and data[f] not in ("lipnik","praha"):
                    return jsonify({"ok": False, "error":"bad_site"}), 400
                if f=="category" and data[f] not in VALID_CATS:
                    return jsonify({"ok": False, "error":"bad_category"}), 400
                updates.append(f"{f}=?")
                params.append(float(data[f]) if f=="qty" else data[f])
        if not updates: return jsonify({"ok": False, "error":"nothing_to_update"}), 400
        params.append(iid)
        db.execute(f"UPDATE items SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        iid = request.args.get("id", type=int)
        if not iid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM items WHERE id=?", (iid,))
        db.commit()
        return jsonify({"ok": True})

# ---------- jobs ----------
@app.route("/api/jobs", methods=["GET","POST","PATCH","DELETE"])
def api_jobs():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()
    if request.method == "GET":
        rows = db.execute("SELECT * FROM jobs ORDER BY date DESC, id DESC").fetchall()
        return jsonify({"ok": True, "jobs":[dict(r) for r in rows]})
    data = request.get_json(force=True, silent=True) or {}
    if request.method == "POST":
        req = ["title","client","city","code","date"]
        if not all(data.get(k) for k in req):
            return jsonify({"ok": False, "error":"invalid_input"}), 400
        status = data.get("status") or "Plán"
        note = data.get("note") or ""
        db.execute("""INSERT INTO jobs(title,client,status,city,code,date,note)
                      VALUES (?,?,?,?,?,?,?)""",
                   (data["title"], data["client"], status, data["city"], data["code"], data["date"], note))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "PATCH":
        jid = data.get("id")
        if not jid:
            return jsonify({"ok": False, "error":"missing_id"}), 400
        fields = ["title","client","status","city","code","date","note"]
        updates = []; params = []
        for f in fields:
            if f in data:
                updates.append(f"{f}=?"); params.append(data[f])
        if not updates: return jsonify({"ok": False, "error":"nothing_to_update"}), 400
        params.append(jid)
        db.execute(f"UPDATE jobs SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        jid = request.args.get("id", type=int)
        if not jid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM jobs WHERE id=?", (jid,))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/jobs/<int:jid>", methods=["GET"])
def api_job_detail(jid):
    u, err = require_auth()
    if err: return err
    db = get_db()
    job = db.execute("SELECT * FROM jobs WHERE id=?", (jid,)).fetchone()
    if not job: return jsonify({"ok": False, "error":"not_found"}), 404
    mats = [dict(r) for r in db.execute("SELECT * FROM job_materials WHERE job_id=? ORDER BY id DESC", (jid,)).fetchall()]
    tools = [dict(r) for r in db.execute("SELECT * FROM job_tools WHERE job_id=? ORDER BY id DESC", (jid,)).fetchall()]
    photos = [dict(r) for r in db.execute("SELECT * FROM job_photos WHERE job_id=? ORDER BY id DESC", (jid,)).fetchall()]
    assigns = [r["employee_id"] for r in db.execute("SELECT employee_id FROM job_assignments WHERE job_id=?", (jid,)).fetchall()]
    tasks = [dict(r) for r in db.execute("SELECT * FROM tasks WHERE job_id=? ORDER BY (due_date IS NULL), due_date ASC, id DESC", (jid,)).fetchall()]
    hours = [dict(r) for r in db.execute("""SELECT t.id,t.employee_id,e.name as employee_name,t.date,t.hours,t.place,t.activity
                                           FROM timesheets t LEFT JOIN employees e ON e.id=t.employee_id
                                           WHERE t.job_id=? ORDER BY t.date DESC, t.id DESC""",(jid,)).fetchall()]
    return jsonify({"ok": True, "job": dict(job), "materials": mats, "tools": tools, "photos": photos, "assignments": assigns, "tasks": tasks, "hours": hours})

@app.route("/api/jobs/<int:jid>/materials", methods=["POST","DELETE"])
def api_job_materials(jid):
    u, err = require_role(write=True)
    if err: return err
    db = get_db()
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        name = data.get("name"); qty = data.get("qty"); unit = data.get("unit")
        if not (name and unit and qty is not None): return jsonify({"ok": False, "error":"invalid_input"}), 400
        db.execute("INSERT INTO job_materials(job_id,name,qty,unit) VALUES (?,?,?,?)", (jid, name, float(qty), unit))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        mid = request.args.get("id", type=int)
        if not mid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM job_materials WHERE id=? AND job_id=?", (mid, jid))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/jobs/<int:jid>/tools", methods=["POST","DELETE"])
def api_job_tools(jid):
    u, err = require_role(write=True)
    if err: return err
    db = get_db()
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        name = data.get("name"); qty = data.get("qty"); unit = data.get("unit")
        if not (name and unit and qty is not None): return jsonify({"ok": False, "error":"invalid_input"}), 400
        db.execute("INSERT INTO job_tools(job_id,name,qty,unit) VALUES (?,?,?,?)", (jid, name, float(qty), unit))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        tid = request.args.get("id", type=int)
        if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM job_tools WHERE id=? AND job_id=?", (tid, jid))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/jobs/<int:jid>/photos", methods=["POST","DELETE"])
def api_job_photos(jid):
    u, err = require_role(write=True)
    if err: return err
    db = get_db()
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        data_url = data.get("data_url")
        if not data_url or not data_url.startswith("data:image/"):
            return jsonify({"ok": False, "error":"invalid_image"}), 400
        m = re.match(r"data:image/(png|jpg|jpeg);base64,(.+)", data_url)
        if not m: return jsonify({"ok": False, "error":"invalid_format"}), 400
        ext = "jpg" if m.group(1)=="jpeg" else m.group(1)
        raw = base64.b64decode(m.group(2))
        fname = f"job{jid}-{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}.{ext}"
        with open(os.path.join(UPLOAD_DIR, fname), "wb") as f:
            f.write(raw)
        db.execute("INSERT INTO job_photos(job_id,filename,created_at) VALUES (?,?,?)", (jid, fname, datetime.utcnow().isoformat()))
        db.commit()
        return jsonify({"ok": True, "filename": fname})
    if request.method == "DELETE":
        pid = request.args.get("id", type=int)
        if not pid: return jsonify({"ok": False, "error":"missing_id"}), 400
        row = db.execute("SELECT filename FROM job_photos WHERE id=? AND job_id=?", (pid, jid)).fetchone()
        if row:
            try: os.remove(os.path.join(UPLOAD_DIR, row["filename"]))
            except Exception: pass
        db.execute("DELETE FROM job_photos WHERE id=? AND job_id=?", (pid, jid))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/jobs/<int:jid>/assignments", methods=["GET","POST"])
def api_job_assignments(jid):
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()
    if request.method == "GET":
        rows = db.execute("SELECT employee_id FROM job_assignments WHERE job_id=?", (jid,)).fetchall()
        return jsonify({"ok": True, "employee_ids":[r["employee_id"] for r in rows]})
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get("employee_ids") or []
    if not isinstance(ids, list): return jsonify({"ok": False, "error":"invalid_input"}), 400
    db.execute("DELETE FROM job_assignments WHERE job_id=?", (jid,))
    for eid in ids:
        try:
            db.execute("INSERT INTO job_assignments(job_id,employee_id) VALUES (?,?)", (jid, int(eid)))
        except Exception:
            pass
    db.commit()
    return jsonify({"ok": True})

# ---------- tasks ----------
VALID_STATUS = ('nový','probíhá','hotovo')

@app.route("/api/tasks", methods=["GET","POST","PATCH","DELETE"])
def api_tasks():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()
    if request.method == "GET":
        mine = request.args.get("mine")
        job_id = request.args.get("job_id", type=int)
        q = "SELECT * FROM tasks"
        conds = []; params = []
        if mine and u:
            conds.append("employee_id=?"); params.append(u["id"])
        if job_id:
            conds.append("job_id=?"); params.append(job_id)
        if conds:
            q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY (due_date IS NULL), due_date ASC, id DESC"
        rows = db.execute(q, params).fetchall()
        return jsonify({"ok": True, "tasks":[dict(r) for r in rows]})
    data = request.get_json(force=True, silent=True) or {}
    if request.method == "POST":
        title = data.get("title"); description = data.get("description") or ""
        due_date = data.get("due_date"); status = data.get("status") or "nový"
        employee_id = data.get("employee_id"); job_id = data.get("job_id")
        if status not in VALID_STATUS:
            return jsonify({"ok": False, "error":"bad_status"}), 400
        if not title:
            return jsonify({"ok": False, "error":"missing_title"}), 400
        db.execute("""INSERT INTO tasks(title,description,status,due_date,employee_id,job_id,created_by,created_at)
                      VALUES (?,?,?,?,?,?,?,?)""",
                   (title, description, status, due_date, employee_id, job_id, u["id"], datetime.utcnow().isoformat()))
        db.commit()
        return jsonify({"ok": True})
    if request.method == "PATCH":
        tid = data.get("id")
        if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
        updates=[]; params=[]
        for f in ["title","description","status","due_date","employee_id","job_id"]:
            if f in data:
                if f=="status" and data[f] not in VALID_STATUS:
                    return jsonify({"ok": False, "error":"bad_status"}), 400
                updates.append(f"{f}=?"); params.append(data[f])
        if not updates: return jsonify({"ok": False, "error":"nothing_to_update"}), 400
        params.append(tid)
        db.execute(f"UPDATE tasks SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
        return jsonify({"ok": True})
    if request.method == "DELETE":
        tid = request.args.get("id", type=int)
        if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM tasks WHERE id=?", (tid,))
        db.commit()
        return jsonify({"ok": True})

# ---------- exports ----------
@app.route("/export/employee_hours.xlsx")
def export_employee_hours():
    u, err = require_auth()
    if err: return err
    emp = request.args.get("employee_id", type=int)
    if not emp: return jsonify({"ok": False, "error":"missing_employee"}), 400
    db = get_db()
    rows = db.execute("""SELECT t.date,t.hours,t.place,t.activity,j.title,j.code
                         FROM timesheets t JOIN jobs j ON j.id=t.job_id
                         WHERE t.employee_id=? ORDER BY t.date ASC""",(emp,)).fetchall()
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"ok": False, "error":"openpyxl_missing"}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hodiny"
    ws.append(["Datum","Hodiny","Zakázka","Kód","Místo","Popis činnosti"])
    for r in rows:
        ws.append([r["date"], float(r["hours"]), r["title"], r["code"], r["place"] or "", r["activity"] or ""])
    for i,w in enumerate([14,10,40,16,20,40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="employee_hours.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export/job_materials.xlsx")
def export_job_materials():
    u, err = require_auth()
    if err: return err
    jid = request.args.get("job_id", type=int)
    if not jid: return jsonify({"ok": False, "error":"missing_job"}), 400
    db = get_db()
    job = db.execute("SELECT title,code FROM jobs WHERE id=?", (jid,)).fetchone()
    mats = db.execute("SELECT name,qty,unit FROM job_materials WHERE job_id=? ORDER BY id ASC",(jid,)).fetchall()
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"ok": False, "error":"openpyxl_missing"}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiál"
    ws.append(["Zakázka", job["title"] if job else "", "Kód", job["code"] if job else ""])
    ws.append([])
    ws.append(["Název","Množství","Jednotka"])
    for r in mats:
        ws.append([r["name"], float(r["qty"]), r["unit"]])
    for i,w in enumerate([40,12,12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="job_materials.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export/warehouse.xlsx")
def export_warehouse():
    u, err = require_auth()
    if err: return err
    db = get_db()
    rows = db.execute("SELECT site,category,name,qty,unit FROM items ORDER BY site,category,name").fetchall()
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"ok": False, "error":"openpyxl_missing"}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sklad"
    ws.append(["Stanoviště","Kategorie","Název","Množství","Jednotka"])
    for r in rows:
        ws.append([r["site"], r["category"], r["name"], float(r["qty"]), r["unit"]])
    for i,w in enumerate([14,22,50,12,12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="warehouse.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
